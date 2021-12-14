import requests 
import json
from openpyxl import load_workbook

def read_table(sheet, columnnames, header_row = 1):
    name_to_index = {
        n: i for i, n
        in enumerate(c.value for c in sheet[header_row])
        if n is not None}
    column_indices = [name_to_index[n] for n in columnnames]
    print(column_indices)
    for i in range(header_row + 1, ws.max_row): #
        yield dict(zip(columnnames, (ws.cell(row = i, column = x + 1).value for x in column_indices)))
 

def get_google_results(address, api_key=None, return_full_response=False):
    """
    Get geocode results from Google Maps Geocoding API.
    
    Note, that in the case of multiple google geocode reuslts, this function returns details of the FIRST result.
    
    @param address: String address as accurate as possible. For Example "18 Grafton Street, Dublin, Ireland"
    @param api_key: String API key if present from google. 
                    If supplied, requests will use your allowance from the Google API. If not, you
                    will be limited to the free usage of 2500 requests per day.
    @param return_full_response: Boolean to indicate if you'd like to return the full response from google. This
                    is useful if you'd like additional location details for storage or parsing later.
    """
    # Set up your Geocoding url
    geocode_url = "https://maps.googleapis.com/maps/api/geocode/json?address={}".format(address)
    if api_key is not None:
        geocode_url = geocode_url + "&key={}".format(api_key)
        
    # Ping google for the reuslts:
    results = requests.get(geocode_url)
    # Results will be in JSON format - convert to dict using requests functionality
    results = results.json()
    
    # if there's no results or an error, return empty results.
    if len(results['results']) == 0:
        output = {
            "formatted_address" : None,
            "latitude": None,
            "longitude": None,
            "accuracy": None,
            "google_place_id": None,
            "type": None,
            "postcode": None
        }
    else:    
        answer = results['results'][0]
        output = {
            "formatted_address" : answer.get('formatted_address'),
            "latitude": answer.get('geometry').get('location').get('lat'),
            "longitude": answer.get('geometry').get('location').get('lng'),
            "accuracy": answer.get('geometry').get('location_type'),
            "google_place_id": answer.get("place_id"),
            "type": ",".join(answer.get('types')),
            "postcode": ",".join([x['long_name'] for x in answer.get('address_components') 
                                  if 'postal_code' in x.get('types')])
        }
        
    # Append some other details:    
    output['input_string'] = address
    output['number_of_results'] = len(results['results'])
    output['status'] = results.get('status')
    if return_full_response is True:
        output['response'] = results
    
    return output

def write_json_file(filename, rows):
    with open(filename, 'a') as f:
        json.dump(rows, f)
    return

api_key = 'xxxx'
opfile = 'xxxx.json'
wb = load_workbook(filename = './data/xxx.xlsx', read_only=True)
ws = wb.worksheets[0]
cols = ['CODE']
#col_indices = {n for n, cell in enumerate(ws.rows[0]) if cell.value in cols}
column_list = [cell.column for cell in ws[1]]
print(column_list)

rows = read_table(ws, 'CODE Name Type AddressLine1 CityName StateCode ZipCode'.split())
for row_dict in rows:
    if not ('' in row_dict.items()):
       addr = '{0},{1},{2},{3}'.format(row_dict['AddressLine1'], row_dict['CityName'], row_dict['StateCode'], row_dict['ZipCode'])        
       geo_res = get_google_results(addr, api_key)
       #row_dict['geo_res'] = get_google_results(addr, api_key)      
       row_transform = {
           'id': row_dict['CODE'],
           'name': row_dict['Name'],
           'ty': row_dict['Type'][:1],
           'lat': geo_res.get('latitude'),
           'lon': geo_res.get('longitude')
       }
       write_json_file(opfile, row_transform)
       #print(row_dict)
print('completed')
 