import requests 
import json
import itertools 
from operator import itemgetter
from openpyxl import load_workbook

def column_2_indices(sheet, columnnames, header_row = 1):
    name_to_index = {
        n: i for i, n
        in enumerate(c.value for c in sheet[header_row])
        if n is not None}
    column_indices = [name_to_index[n] for n in columnnames]
    return column_indices 

def read_rows(sheet, keynames, column_indices, offset, size):
    return  [list(zip(keynames, (sheet.cell(row = i, column = x + 1).value for x in column_indices))) for i in range(offset, offset + size)]

def read_table_rows(sheet, columnnames, column_indices, offset = 1, size = 500):    
    keynames = [c.lower() for c in columnnames]
    print(column_indices)
    return [list(zip(keynames, (sheet.cell(row = i, column = x + 1).value for x in column_indices))) for i in range(offset, size)]

 

def write_json_file(filename, rows):
    with open(filename, 'a') as f:
        json.dump(rows, f)
    return

def open_sheet(filename, sheetname):
    wb = load_workbook(filename = filename, read_only=True)
    return wb.get_sheet_by_name(sheetname)


opfilename = 'C:/Users/xxxx/Downloads/data/{0}.json'
wb = load_workbook(filename = './data/xxxxs.xlsx', read_only=True)
ws = wb.worksheets[0]
cols = ['Year','Id', 'Name', 'City', 'County', 'GradeServed']
colindices = column_2_indices(ws, cols)
keynames = [c.lower() for c in cols]
offset = 2
size = 500
while offset <=  ws.max_row:
    print(offset)
    for year, entities in itertools.groupby(read_rows(ws, keynames, colindices, offset + 1, size), key=lambda x: x[0]):
        jfile = opfilename.format(year[1])  
        for row_dict in entities:         
            write_json_file(jfile, dict((k, v) for k, v in row_dict if k != 'year'))      
    offset += size

# for year, entities in itertools.groupby(read_table(ws, 'Type Year Id Name City County'.split()), key=lambda x: x[1]):
#     jfile = opfilename.format(year[1])  
#     for row_dict in entities:         
#         write_json_file(jfile, dict((k, v) for k, v in row_dict))
       #print(row_dict)
print('completed')